"""
Parser for XLSX documents.
"""

import logging
import os
from typing import List, Dict, Any
import openpyxl

from src.parsers.base import BaseParser

logger = logging.getLogger(__name__)


class XLSXParser(BaseParser):
    """Parser for XLSX documents."""
    
    @staticmethod
    def supported_mime_types() -> List[str]:
        """
        Get the list of MIME types supported by this parser.
        
        Returns:
            List of supported MIME types
        """
        return [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/vnd.ms-excel'
        ]
    
    @classmethod
    def can_parse(cls, mime_type: str) -> bool:
        """
        Check if this parser can handle the given mime type.
        
        Args:
            mime_type: MIME type of the document
            
        Returns:
            True if this parser can handle the mime type, False otherwise
        """
        return mime_type in cls.supported_mime_types() or 'excel' in mime_type.lower()
    
    def parse(self, file_path: str) -> str:
        """
        Parse an XLSX file and extract text content.
        
        Args:
            file_path: Path to the XLSX file
            
        Returns:
            Extracted text content
        """
        if not os.path.exists(file_path):
            logger.error(f"XLSX file not found: {file_path}")
            return ""
        
        try:
            text_content = []
            
            # Open the workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Extract properties if available
            if hasattr(workbook, 'properties'):
                props = workbook.properties
                if hasattr(props, 'title') and props.title:
                    text_content.append(f"Title: {props.title}")
                if hasattr(props, 'creator') and props.creator:
                    text_content.append(f"Author: {props.creator}")
                text_content.append("")  # Empty line
            
            # Process each worksheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_content.append(f"--- Sheet: {sheet_name} ---")
                
                # Get rows and columns
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                # Only process if there's actual data
                if max_row > 0 and max_col > 0:
                    # Extract column headers (first row)
                    headers = []
                    for col in range(1, max_col + 1):
                        cell = sheet.cell(row=1, column=col)
                        headers.append(str(cell.value) if cell.value is not None else "")
                    
                    if any(headers):  # Only append if there are actual headers
                        text_content.append(" | ".join(headers))
                        text_content.append("-" * (sum(len(h) for h in headers) + 3 * (len(headers) - 1)))
                    
                    # Extract data rows (limit to reasonable number to avoid huge outputs)
                    max_rows_to_extract = min(max_row, 200)  # Limit to 200 rows
                    for row in range(2, max_rows_to_extract + 1):
                        row_data = []
                        for col in range(1, max_col + 1):
                            cell = sheet.cell(row=row, column=col)
                            row_data.append(str(cell.value) if cell.value is not None else "")
                        
                        if any(row_data):  # Only append non-empty rows
                            text_content.append(" | ".join(row_data))
                    
                    if max_row > max_rows_to_extract:
                        text_content.append(f"... (truncated, {max_row - max_rows_to_extract} more rows)")
                
                # Add an empty line between sheets
                text_content.append("")
            
            return "\n".join(text_content)
            
        except Exception as e:
            logger.exception(f"Error parsing XLSX file {file_path}: {e}")
            return f"Error parsing XLSX file: {str(e)}"
